#!/usr/bin/python3
"""
Script to extract the solution worksheet from Finalreal.xlsx and save it as solution.xlsx
"""

import os
import sys
import openpyxl
from pathlib import Path

def extract_solution(source_path, destination_path):
    """
    Extract the solution worksheet from the source file and save it to the destination file
    
    Args:
        source_path: Path to the source Excel file (Finalreal.xlsx)
        destination_path: Path where the solution Excel file will be saved
    """
    try:
        print(f"Loading source file: {source_path}")
        source_wb = openpyxl.load_workbook(source_path)
        
        if "solution" not in source_wb.sheetnames:
            print(f"Error: 'solution' worksheet not found in {source_path}")
            print(f"Available worksheets: {source_wb.sheetnames}")
            return False
        
        # Create a new workbook for the solution
        dest_wb = openpyxl.Workbook()
        
        # Remove the default sheet
        if "Sheet" in dest_wb.sheetnames:
            default_sheet = dest_wb["Sheet"]
            dest_wb.remove(default_sheet)
        
        # Create both blank and solution sheets to match the expected structure
        blank_sheet = dest_wb.create_sheet("blank")
        solution_sheet = dest_wb.create_sheet("solution")
        
        # Copy data from source solution sheet to destination solution sheet
        source_solution = source_wb["solution"]
        
        # Copy row 1 (Y/N values) and other essential rows
        for row_idx in range(1, min(10, source_solution.max_row + 1)):
            for col_idx in range(1, source_solution.max_column + 1):
                cell_value = source_solution.cell(row=row_idx, column=col_idx).value
                solution_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # For row 1, also create blank sheet with mostly N values
                if row_idx == 1 and col_idx >= 5:  # Column E onwards
                    # Set blank sheet to have N values except for a couple Y values to simulate partial correctness
                    if col_idx < 7:  # First two cells as Y for testing
                        blank_sheet.cell(row=row_idx, column=col_idx, value="Y")
                    else:
                        blank_sheet.cell(row=row_idx, column=col_idx, value="N")
                elif row_idx == 1:
                    # Copy A1-D1 values directly
                    blank_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Add formulas for cells B1 and D1 in both sheets
        solution_sheet["B1"] = '=SUM(IF(E1:CA1="Y",1,0))'
        solution_sheet["D1"] = '=SUM(IF(E1:CA1="N",1,0))'
        blank_sheet["B1"] = '=SUM(IF(E1:CA1="Y",1,0))'
        blank_sheet["D1"] = '=SUM(IF(E1:CA1="N",1,0))'
        
        # Save the solution file
        dest_wb.save(destination_path)
        print(f"Solution file saved to: {destination_path}")
        
        # Also create a sample student submission for testing
        sample_path = os.path.join(os.path.dirname(destination_path), "sample_submission.xlsx")
        dest_wb.save(sample_path)
        print(f"Sample submission file saved to: {sample_path}")
        
        return True
    
    except Exception as e:
        print(f"Error extracting solution: {e}")
        return False

def main():
    # Get the source and destination paths
    if len(sys.argv) < 2:
        print("Usage: python extract_solution.py [path/to/Finalreal.xlsx]")
        print("If no path is provided, will look for Finalreal.xlsx in the current directory")
        source_path = "Finalreal.xlsx"
    else:
        source_path = sys.argv[1]
    
    # Check if source file exists
    if not os.path.exists(source_path):
        print(f"Error: Source file '{source_path}' not found")
        return 1
    
    # Set destination path
    script_dir = os.path.dirname(os.path.abspath(__file__))
    destination_path = os.path.join(script_dir, "solution.xlsx")
    
    # Extract the solution
    if extract_solution(source_path, destination_path):
        print("\nSolution extraction completed successfully")
        print("\nNext steps:")
        print("1. Review the solution.xlsx file to ensure it has the correct Y/N values")
        print("2. Build the Docker image with: docker build -t excel-autograder .")
        print("3. Test the autograder with the sample_submission.xlsx file")
        return 0
    else:
        print("\nSolution extraction failed")
        return 1

if __name__ == "__main__":
    sys.exit(main())