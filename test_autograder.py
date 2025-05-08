#!/usr/bin/python3
"""
Script to test the Excel autograder locally without using Docker
"""

import os
import sys
import json
import tempfile
import openpyxl
from pathlib import Path

# Import the autograder function
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from excel_autograder import grade_excel_worksheet

def create_test_file(y_count, n_count):
    """
    Create a test Excel file with specified Y/N counts
    
    Args:
        y_count: Number of Y values to include
        n_count: Number of N values to include
        
    Returns:
        Path to the created file
    """
    wb = openpyxl.Workbook()
    
    # Create both sheets
    if "blank" not in wb.sheetnames:
        wb.create_sheet("blank")
    if "solution" not in wb.sheetnames:
        wb.create_sheet("solution")
    
    # Remove the default sheet if it exists
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheet = wb["blank"]
    
    # Add Y: and N: headers
    sheet["A1"] = "Y:"
    sheet["B1"] = "=SUM(IF(E1:CA1=\"Y\",1,0))"
    sheet["C1"] = "N:"
    sheet["D1"] = "=SUM(IF(E1:CA1=\"N\",1,0))"
    
    # Add Y/N values starting from column E
    for i in range(y_count):
        col = openpyxl.utils.get_column_letter(i + 5)  # Start from column E (5)
        sheet[f"{col}1"] = "Y"
    
    for i in range(y_count, y_count + n_count):
        col = openpyxl.utils.get_column_letter(i + 5)  # Continue after Y values
        sheet[f"{col}1"] = "N"
    
    # Save to a temporary file
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    
    return path

def test_with_sample():
    """Test the autograder with the sample_submission.xlsx file"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    sample_path = os.path.join(script_dir, "sample_submission.xlsx")
    solution_path = os.path.join(script_dir, "solution.xlsx")
    
    if not os.path.exists(sample_path):
        print(f"Error: Sample file '{sample_path}' not found")
        print("Run extract_solution.py first to create the sample file")
        return False
    
    if not os.path.exists(solution_path):
        print(f"Error: Solution file '{solution_path}' not found")
        print("Run extract_solution.py first to create the solution file")
        return False
    
    print("Testing with sample_submission.xlsx...")
    result = grade_excel_worksheet(sample_path, solution_path)
    
    print("\nGrading Result:")
    print(f"Score: {result['score']:.2%}")
    print(f"Feedback:\n{result['feedback']}")
    
    return True

def test_with_generated_files():
    """Test the autograder with generated test files"""
    print("Creating test files...")
    
    # Create solution file with all Y values
    solution_file = create_test_file(20, 0)  # 20 Y values, 0 N values
    print(f"Solution file created: {solution_file}")
    
    # Create student file with some Y and some N values
    student_file = create_test_file(5, 15)  # 5 Y values, 15 N values
    print(f"Student file created: {student_file}")
    
    print("\nTesting with generated files...")
    result = grade_excel_worksheet(student_file, solution_file)
    
    print("\nGrading Result:")
    print(f"Score: {result['score']:.2%}")
    print(f"Feedback:\n{result['feedback']}")
    
    # Clean up
    try:
        os.remove(solution_file)
        os.remove(student_file)
    except:
        pass
    
    return True

def main():
    """Run tests for the Excel autograder"""
    print("=== Excel Autograder Test ===\n")
    
    # First try with sample file
    if os.path.exists("sample_submission.xlsx") and os.path.exists("solution.xlsx"):
        test_with_sample()
    else:
        # If sample and solution files don't exist, use generated files
        test_with_generated_files()
    
    print("\nTest completed!")
    return 0

if __name__ == "__main__":
    sys.exit(main())