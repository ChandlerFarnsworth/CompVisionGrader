#!/usr/bin/python3

import pandas as pd
import numpy as np
import json
import os
import sys
import shutil
from pathlib import Path
import openpyxl
import re

# Constants
SUBMISSION_DIR = "submissions"
SOLUTION_DIR = "solutions"
OUTPUT_DIR = "output"
SUBMISSION_FILE = "submission.xlsx"
SOLUTION_FILE = "solution.xlsx"

def print_stderr(message):
    """Print a message to stderr."""
    print(message, file=sys.stderr)

def send_feedback(score, feedback):
    """Send feedback to the grading system."""
    output = {
        "fractionalScore": score,
        "feedback": feedback
    }
    print(json.dumps(output))
    
    # Also save to a file
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(os.path.join(OUTPUT_DIR, "feedback.json"), "w") as f:
        json.dump(output, f, indent=2)

def clean_value(val):
    """Clean a cell value by removing currency symbols and whitespace."""
    if val is None or val == "":
        return ""
    if isinstance(val, str):
        val = re.sub(r"[\$,]", "", val).strip()
    return val

def compare_cells(student_val, solution_val):
    """Compare two cell values with numeric tolerance or string equality."""
    try:
        return np.isclose(float(student_val), float(solution_val), atol=0.01)
    except:
        return str(student_val).strip() == str(solution_val).strip()

def safe_to_numeric(value):
    """Safely convert value to numeric, handling errors explicitly to avoid deprecation warning."""
    if pd.isna(value) or value == "":
        return value
    
    try:
        return float(value)
    except (ValueError, TypeError):
        return value

def grade_gan(student_sheet, solution_sheet):
    """Grade the GAN sheet of the workbook."""
    gan_ranges = [
        "B11:U15", "L19:U23", "B29:I30", "L29:U30",
        "AF85:AP95", "AT85:BD95", "AF108:AP118", "AT108:BD118",
        "L125:AA125", "D128:N138", "L144:AA144",
        "L172:AA172", "L174:AA174", "L176:AA176", "L178:L178",
        "L181:AA181", "L183:AA183", "L185:AA185", "L187:L187"
    ]

    correct = 0
    total = 0

    for cell_range in gan_ranges:
        try:
            for row_student, row_solution in zip(student_sheet[cell_range], solution_sheet[cell_range]):
                for cell_student, cell_solution in zip(row_student, row_solution):
                    val_student = clean_value(cell_student.value)
                    val_solution = clean_value(cell_solution.value)
                    if val_solution != "":
                        total += 1
                        if compare_cells(val_student, val_solution):
                            correct += 1
        except Exception as e:
            print_stderr(f"Error processing GAN range {cell_range}: {e}")
            continue

    return correct, total

def get_bordered_cells(sheet):
    """Get all cells with borders in a sheet."""
    cells = []
    for row in sheet.iter_rows():
        for cell in row:
            b = cell.border
            if b and any([b.top.style, b.bottom.style, b.left.style, b.right.style]):
                cells.append(cell.coordinate)
    return cells

def grade_unet(student_sheet, solution_sheet):
    """Grade the U-Net sheet of the workbook based on cells with borders."""
    coords = get_bordered_cells(solution_sheet)
    correct = 0
    total = 0

    for coord in coords:
        try:
            val_student = clean_value(student_sheet[coord].value)
            val_solution = clean_value(solution_sheet[coord].value)
            if val_solution != "":
                total += 1
                if compare_cells(val_student, val_solution):
                    correct += 1
        except Exception as e:
            print_stderr(f"Error processing U-Net cell {coord}: {e}")
            continue

    return correct, total

def grade_classical(student_sheet, solution_sheet):
    """Grade the Classical sheet of the workbook."""
    # Define specific regions that need to be graded in the Classical sheet
    classical_ranges = [
        # Cosine Similarity section
        "K4:K6", "P4:P4", "S4:S5", "W4:W5", "AB4:AB4",
        
        # Transformations sections
        "M53:N56", "Y53:Z56", "AC53:AD56",
        
        # Image Gradients section
        "F110:J114", "N110:R114", "V110:Z114", "AE110:AI114"
    ]
    
    correct = 0
    total = 0
    
    for cell_range in classical_ranges:
        try:
            for row_student, row_solution in zip(student_sheet[cell_range], solution_sheet[cell_range]):
                for cell_student, cell_solution in zip(row_student, row_solution):
                    val_student = clean_value(cell_student.value)
                    val_solution = clean_value(cell_solution.value)
                    if val_solution != "":
                        total += 1
                        if compare_cells(val_student, val_solution):
                            correct += 1
        except Exception as e:
            print_stderr(f"Error processing Classical range {cell_range}: {e}")
            continue
    
    # Additionally, check for specific individual cells that contain important calculations
    individual_cells = [
        "K21", "K22", "K23", "C25", "G25", "K25", "O25", "S25", "W25", "AB25",
        "C37", "G37", "K37", "O37", "S37", "W37", "AB37",
        "C49", "G49"
    ]
    
    for coord in individual_cells:
        try:
            val_student = clean_value(student_sheet[coord].value)
            val_solution = clean_value(solution_sheet[coord].value)
            if val_solution != "":
                total += 1
                if compare_cells(val_student, val_solution):
                    correct += 1
        except Exception as e:
            print_stderr(f"Error processing Classical cell {coord}: {e}")
            continue
    
    return correct, total

def find_submission_file(submission_dir=SUBMISSION_DIR):
    """Find the Excel file in the submission directory."""
    if not os.path.exists(submission_dir):
        return None
    
    for filename in os.listdir(submission_dir):
        if filename.endswith(('.xlsx', '.xls')):
            return os.path.join(submission_dir, filename)
    
    return None

def copy_submission_file(source_dir, destination_dir=None):
    """Copy the submission file to a new location."""
    if destination_dir is None:
        destination_dir = SUBMISSION_DIR
        
    # Ensure the destination directory exists
    os.makedirs(destination_dir, exist_ok=True)
    
    # Find any Excel file in the source directory
    excel_file = None
    for f in os.listdir(source_dir):
        if f.endswith(('.xlsx', '.xls')):
            excel_file = f
            break
    
    if excel_file is None:
        return None
    
    # Copy the file
    source_path = os.path.join(source_dir, excel_file)
    dest_path = os.path.join(destination_dir, SUBMISSION_FILE)
    shutil.copyfile(source_path, dest_path)
    
    return dest_path

def grade_workbook(submission_path, solution_path):
    """Grade a submission against a solution."""
    try:
        # Load workbooks with data_only=True to get calculated values
        wb_student = openpyxl.load_workbook(submission_path, data_only=True)
        wb_solution = openpyxl.load_workbook(solution_path, data_only=True)
        
        # Check if all required sheets exist
        required_sheets = ["Classical", "GAN", "U-Net"]
        missing_sheets = []
        
        for sheet in required_sheets:
            if sheet not in wb_student.sheetnames:
                missing_sheets.append(sheet)
        
        if missing_sheets:
            return 0.0, f"Missing sheets in submission: {', '.join(missing_sheets)}"
        
        # Grade each sheet
        classical_correct, classical_total = grade_classical(wb_student["Classical"], wb_solution["Classical"])
        gan_correct, gan_total = grade_gan(wb_student["GAN"], wb_solution["GAN"])
        unet_correct, unet_total = grade_unet(wb_student["U-Net"], wb_solution["U-Net"])
        
        # Calculate total score
        total_correct = classical_correct + gan_correct + unet_correct
        total_cells = classical_total + gan_total + unet_total
        
        if total_cells == 0:
            return 0.0, "No cells were evaluated. This may indicate an issue with the submission format."
        
        final_score = round(total_correct / total_cells, 2)
        
        # Calculate percentages safely
        classical_percent = (classical_correct / classical_total * 100) if classical_total > 0 else 0
        gan_percent = (gan_correct / gan_total * 100) if gan_total > 0 else 0
        unet_percent = (unet_correct / unet_total * 100) if unet_total > 0 else 0
        
        # Create feedback
        feedback = f"You scored {total_correct}/{total_cells} correct ({final_score * 100:.0f}%).\n\n"
        feedback += f"Classical: {classical_correct}/{classical_total} ({classical_percent:.0f}%)\n"
        feedback += f"GAN: {gan_correct}/{gan_total} ({gan_percent:.0f}%)\n"
        feedback += f"U-Net: {unet_correct}/{unet_total} ({unet_percent:.0f}%)"
        
        # Add overall assessment
        if final_score >= 0.9:
            feedback += "\n\nExcellent work!"
        elif final_score >= 0.8:
            feedback += "\n\nGood job! A few answers need improvement."
        elif final_score >= 0.7:
            feedback += "\n\nYou're on the right track, but several answers need revision."
        else:
            feedback += "\n\nPlease review your work and try again."
        
        return final_score, feedback
        
    except Exception as e:
        return 0.0, f"Error grading submission: {str(e)}"

def testRunner():
    """Main grading function, similar to the test_runner.py example."""
    # Check for submission location from environment variable
    submission_location = os.environ.get('SUBMISSION_DIR', "/shared/submission")
    
    # Setup paths
    submission_destination = os.path.join(SUBMISSION_DIR, SUBMISSION_FILE)
    solution_path = os.path.join(SOLUTION_DIR, SOLUTION_FILE)
    
    # Ensure directories exist
    os.makedirs(SUBMISSION_DIR, exist_ok=True)
    os.makedirs(SOLUTION_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Check if we need to copy a submission file
    if os.path.exists(submission_location) and os.path.isdir(submission_location):
        # Try to find and copy an Excel file
        submission_file = copy_submission_file(submission_location)
        if not submission_file:
            return {"score": 0.0, "feedback": "Your submission file could not be found. Please ensure you've uploaded an Excel file (.xlsx or .xls)."}
    else:
        # Look for an existing submission file
        submission_file = find_submission_file()
        if not submission_file:
            return {"score": 0.0, "feedback": "No Excel submission file found. Please submit an .xlsx or .xls file."}
    
    # Check if solution file exists
    if not os.path.exists(solution_path):
        return {"score": 0.0, "feedback": "Solution file not found. Contact your instructor."}
    
    # Grade the submission
    score, feedback = grade_workbook(submission_file, solution_path)
    
    return {"score": score, "feedback": feedback}

def main():
    """Main entry point for the script."""
    try:
        # Check if a part ID is provided (similar to your boss's example)
        part_id = os.environ.get('partId')
        if part_id and part_id != "Lg9eS":  # Replace with your actual part ID
            print_stderr("Cannot find matching partId. Please double check your partId's")
            send_feedback(0.0, "Please verify that you have submitted to the proper part of the assignment.")
            return 1
        
        # Run the grader
        result = testRunner()
        send_feedback(result["score"], result["feedback"])
        
        # Return success if score is positive
        return 0 if result["score"] > 0 else 1
        
    except Exception as e:
        print_stderr(f"Error in grader: {str(e)}")
        send_feedback(0.0, f"An error occurred while grading: {str(e)}")
        return 1

if __name__ == "__main__":
    sys.exit(main())