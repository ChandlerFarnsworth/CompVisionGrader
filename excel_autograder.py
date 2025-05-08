#!/usr/bin/python3

import os
import sys
import json
import openpyxl
from pathlib import Path

# Constants
SUBMISSION_LOCATION = "/shared/submission"  # Coursera autograder path
STUDENT_SHEET_NAME = "blank"
SOLUTION_SHEET_NAME = "solution"
PART_ID = "Lg9eS"  # Update this with the correct part ID

def print_stderr(error_msg):
    """Print error message to stderr"""
    print(str(error_msg), file=sys.stderr)

def send_feedback(score, msg):
    """Send feedback to Coursera autograder"""
    post = {'fractionalScore': score, 'feedback': msg}
    print(json.dumps(post))
    
    # Write feedback to file for Coursera
    try:
        with open("/shared/feedback.json", "w") as outfile:
            json.dump(post, outfile)
    except Exception as e:
        print_stderr(f"Error writing feedback: {e}")

def grade_excel_worksheet(student_file_path, solution_file_path):
    """
    Grade Excel worksheet by comparing Y/N values in row 1
    
    Args:
        student_file_path: Path to the student's Excel file
        solution_file_path: Path to the solution Excel file
        
    Returns:
        Dictionary with score and feedback
    """
    try:
        # Load workbooks
        student_wb = openpyxl.load_workbook(student_file_path, data_only=True)
        solution_wb = openpyxl.load_workbook(solution_file_path, data_only=True)
        
        # Verify sheets exist
        if STUDENT_SHEET_NAME not in student_wb.sheetnames:
            return {
                "score": 0.0,
                "feedback": f"Error: Worksheet '{STUDENT_SHEET_NAME}' not found in your submission."
            }
        
        if SOLUTION_SHEET_NAME not in solution_wb.sheetnames:
            return {
                "score": 0.0,
                "feedback": f"Error: Internal error - Solution worksheet not found."
            }
        
        student_sheet = student_wb[STUDENT_SHEET_NAME]
        solution_sheet = solution_wb[SOLUTION_SHEET_NAME]
        
        # Get Y/N values from row 1 in both sheets
        student_values = []
        solution_values = []
        
        # Get max column to analyze
        max_col = max(solution_sheet.max_column, student_sheet.max_column)
        
        # Start from column E (index 5 in openpyxl)
        for col_idx in range(5, max_col + 1):
            student_cell = student_sheet.cell(row=1, column=col_idx)
            solution_cell = solution_sheet.cell(row=1, column=col_idx)
            
            # Add to values list if not None
            if student_cell.value is not None and solution_cell.value is not None:
                student_values.append({
                    "col": openpyxl.utils.get_column_letter(col_idx),
                    "value": student_cell.value
                })
                solution_values.append({
                    "col": openpyxl.utils.get_column_letter(col_idx),
                    "value": solution_cell.value
                })
        
        # Calculate matches
        matches = 0
        total_cells = len(student_values)
        correct_cells = []
        incorrect_cells = []
        
        for i in range(total_cells):
            student_val = student_values[i]
            solution_val = solution_values[i]
            
            if student_val["value"] == solution_val["value"]:
                matches += 1
                correct_cells.append(student_val["col"] + "1")
            else:
                incorrect_cells.append({
                    "cell": student_val["col"] + "1",
                    "student": student_val["value"],
                    "solution": solution_val["value"]
                })
        
        # Calculate score (as a decimal between 0.0 and 1.0)
        score = matches / total_cells if total_cells > 0 else 0.0
        
        # Generate feedback
        feedback = generate_feedback(matches, total_cells, correct_cells, incorrect_cells)
        
        return {
            "score": score,
            "feedback": feedback
        }
        
    except Exception as e:
        return {
            "score": 0.0,
            "feedback": f"Error grading your submission: {str(e)}"
        }

def generate_feedback(matches, total_cells, correct_cells, incorrect_cells):
    """Generate feedback text based on grading results"""
    # Calculate percentage score
    percentage = (matches / total_cells) * 100 if total_cells > 0 else 0
    
    # Start with overall summary
    feedback = [
        f"Your score: {percentage:.2f}%",
        f"You correctly matched {matches} out of {total_cells} Y/N values.",
        ""
    ]
    
    # Add details about correct and incorrect cells
    if correct_cells:
        feedback.append("Correct cells:")
        # Group consecutive cells for readability
        feedback.append(", ".join(correct_cells))
        feedback.append("")
    
    if incorrect_cells:
        feedback.append("Incorrect cells:")
        for cell in incorrect_cells:
            feedback.append(f"{cell['cell']}: Your answer was '{cell['student']}', but should be '{cell['solution']}'")
        feedback.append("")
        feedback.append("Please update your worksheet and try again.")
    else:
        feedback.append("Great job! All values are correct.")
    
    return "\n".join(feedback)

def main():
    """Main function for the autograder"""
    try:
        # Get partId from environment (Coursera sets this)
        try:
            part_id = os.environ['partId']
        except KeyError:
            print_stderr("Please provide the partId.")
            send_feedback(0.0, "Please provide the partId.")
            return
        
        # Verify correct partId
        if part_id != PART_ID:
            print_stderr("Cannot find matching partId. Please double check your partId's")
            send_feedback(0.0, "Please verify that you have submitted to the proper part of the assignment.")
            return
        
        # Find the Excel file in the submission directory
        student_file = None
        solution_file = None
        
        # Local testing mode
        if len(sys.argv) > 2 and os.path.exists(sys.argv[1]) and os.path.exists(sys.argv[2]):
            student_file = sys.argv[1]
            solution_file = sys.argv[2]
        else:
            # Coursera mode
            # Find student submission
            for f in os.listdir(SUBMISSION_LOCATION):
                extension = Path(f).suffix.lower()
                if extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
                    student_file = os.path.join(SUBMISSION_LOCATION, f)
                    break
            
            # Get solution file path
            script_dir = os.path.dirname(os.path.abspath(__file__))
            solution_file = os.path.join(script_dir, "solution.xlsx")
        
        # Check if files exist
        if student_file is None:
            send_feedback(0.0, "No Excel file found in your submission. Please submit an Excel file (.xlsx, .xlsm).")
            return
        
        if not os.path.exists(solution_file):
            send_feedback(0.0, "Internal error: Solution file not found.")
            return
        
        # Grade the submission
        result = grade_excel_worksheet(student_file, solution_file)
        
        # Send feedback to Coursera
        send_feedback(result["score"], result["feedback"])
        
    except Exception as e:
        print_stderr(f"Error in autograder: {e}")
        send_feedback(0.0, f"An error occurred while grading your submission: {str(e)}")

if __name__ == "__main__":
    main()